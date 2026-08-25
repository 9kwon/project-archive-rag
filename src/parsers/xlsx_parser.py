"""XLSX 파서: 시트 단위로 표를 추출한다.

성과내역서처럼 표 자체가 내용인 파일이 대상이다.
실험 원시 데이터(수만 행짜리 수치 덩어리)는 검색 대상이 아니므로
행 수 상한을 두고, 넘으면 시트 개요만 남긴다.
"""

from pathlib import Path

from openpyxl import load_workbook

try:
    from .base import Block
except ImportError:  # 스크립트로 직접 실행할 때
    from base import Block

MAX_ROWS = 300  # 이보다 큰 시트는 원시 데이터로 보고 개요만 남긴다


def _row_text(row) -> str:
    cells = ["" if c is None else str(c).strip() for c in row]
    return " | ".join(cells).rstrip(" |")


def parse_xlsx(path: str | Path) -> list[Block]:
    wb = load_workbook(path, data_only=True, read_only=True)
    blocks: list[Block] = []
    idx = 0

    for ws in wb.worksheets:
        n_rows = ws.max_row or 0
        idx += 1
        blocks.append(
            Block(idx, "heading", f"[시트] {ws.title}", locator=ws.title, level=1)
        )

        if n_rows > MAX_ROWS:
            # 원시 데이터 — 헤더 몇 줄만 남겨 "무엇이 들어 있는 시트인지" 알린다
            head = [_row_text(r) for r in ws.iter_rows(max_row=3, values_only=True)]
            idx += 1
            blocks.append(
                Block(
                    idx,
                    "body",
                    f"원시 데이터 시트 ({n_rows:,}행 × {ws.max_column}열). 상단 미리보기:\n"
                    + "\n".join(t for t in head if t),
                    locator=ws.title,
                    heading_path=[ws.title],
                )
            )
            continue

        rows = [_row_text(r) for r in ws.iter_rows(values_only=True)]
        text = "\n".join(r for r in rows if r.strip())
        if text:
            idx += 1
            blocks.append(Block(idx, "table", text, locator=ws.title, heading_path=[ws.title]))

    wb.close()
    return blocks


if __name__ == "__main__":
    import sys

    target = Path(sys.argv[1])
    blocks = parse_xlsx(target)
    print(f"파일 : {target.name}")
    print(f"블록 : {len(blocks)}개, 총 {sum(len(b.text) for b in blocks):,}자")
    print("-" * 70)
    for b in blocks[:6]:
        print(f"[{b.kind:7s}/{b.locator}]")
        for line in b.text.split("\n")[:8]:
            print(f"   {line[:150]}")
        print()
