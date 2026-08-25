"""성과 표를 엑셀로 내보낸다.

해석·집계는 하지 않는다. 원본 표에서 뽑은 값과 그 출처(문서·위치)를 그대로 담아
사람이 대조·편집할 수 있게 하는 것이 목적이다.

    python src/export_excel.py            # 성과정리.xlsx 생성
"""

import re
import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from perf_table import indicator_org
from project import get_profile

OUT = Path("성과정리.xlsx")
_P = get_profile()
YEARS = _P.years

HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
HEAD_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, color="808080")


def _sheet(wb: Workbook, title: str, headers: list[str], widths: list[int], note: str = ""):
    ws = wb.create_sheet(title)
    row = 1
    if note:
        ws.cell(1, 1, note).font = NOTE_FONT
        row = 2
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row, c, h)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row + 1, 1)
    return ws, row + 1


def normalize_indicator(name: str) -> str:
    """지표명을 문서 간 표기 차이에 상관없이 하나로 맞춘다.

    같은 지표가 문서마다 "플랫폼 (측정 정확도)" / "플랫폼(측정 정확도)"
    처럼 적혀 있어, 공백을 정리하지 않으면 별개 지표로 갈린다.
    """
    s = " ".join(name.split())
    s = re.sub(r"\s*\(\s*", "(", s)
    s = re.sub(r"\s*\)", ")", s)
    return s


def sheet_indicators(wb: Workbook, con: sqlite3.Connection):
    """지표별 연차 목표·실적. 계획(목표)과 보고(실적)를 나란히 둔다."""
    headers = ["지표", "단위", "담당(추정)"]
    for y in YEARS:
        headers += [f"{y}차 목표", f"{y}차 실적", f"{y}차 달성률"]
    headers += ["목표 출처", "실적 출처"]
    widths = [40, 8, 10] + [12, 12, 10] * len(YEARS) + [30, 30]

    ws, r = _sheet(
        wb, "성능지표", headers, widths,
        note="달성률은 원본 표의 '달성도(%)' 열이며 목표 대비 달성 여부다(지표 값 아님).",
    )

    data: dict[str, dict] = {}
    for row in con.execute(
        "SELECT indicator, unit, proj_year, target, NULL actual, NULL rate, doc, locator, "
        "'target' src FROM year_target "
        "UNION ALL SELECT indicator, unit, proj_year, year_target, year_actual, rate, doc, "
        "locator, 'actual' FROM year_actual"
    ):
        name = normalize_indicator(row["indicator"])
        e = data.setdefault(name, {"unit": "", "years": {}, "t_src": set(), "a_src": set()})
        if row["unit"] and not e["unit"]:
            e["unit"] = row["unit"]
        y = e["years"].setdefault(row["proj_year"], {})
        if row["target"] and not y.get("target"):
            y["target"] = row["target"]
        if row["actual"]:
            y["actual"] = row["actual"]
        if row["rate"] is not None:
            y["rate"] = row["rate"]
        (e["t_src"] if row["src"] == "target" else e["a_src"]).add(
            f"{row['doc']} {row['locator']}"
        )

    for name in sorted(data, key=lambda n: (indicator_org(n) != _P.home_org, n)):
        e = data[name]
        vals = [name, e["unit"], indicator_org(name)]
        for y in YEARS:
            d = e["years"].get(y, {})
            vals += [d.get("target", ""), d.get("actual", ""), d.get("rate", "")]
        vals += ["; ".join(sorted(e["t_src"])[:2]), "; ".join(sorted(e["a_src"])[:2])]
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c, v)
        r += 1


def sheet_quant(wb: Workbook):
    """정량적 연구개발성과표 — 원본 격자를 그대로 옮긴다.

    이 표는 칸마다 표기 방식이 다르다. "1(기관A)"처럼 건수(기관)인 칸이 있는가 하면
    "자문 보고서(1)아키텍쳐 설계서(1)"처럼 산출물 이름(개수)인 칸도 있다.
    기계적으로 건수를 세면 틀리므로 해석하지 않고 원문을 싣는다.
    """
    from ingest import collect_files, load_config
    from metadata import extract
    from parsers import parse

    ws = wb.create_sheet("정량성과표(원본)")
    ws.cell(1, 1, "원본 표를 해석 없이 그대로 옮긴 것이다. 칸마다 표기 방식이 다르므로 "
                  "건수 집계는 사람이 판단해야 한다.").font = NOTE_FONT
    r = 3

    files = [
        f for f in collect_files(load_config())
        if any(k in f.stem for k in ("연차보고서", "단계보고서"))
    ]
    for f in files:
        doc = extract(f).name
        for b in parse(f):
            if b.kind != "table" or "정량적 연구개발성과표" not in b.text[:200]:
                continue
            cell = ws.cell(r, 1, f"[{doc} · {b.locator}]")
            cell.font = HEAD_FONT
            cell.fill = HEAD_FILL
            r += 1
            for line in b.text.split("\n"):
                for c, v in enumerate(line.split(" | "), start=1):
                    ws.cell(r, c, v)
                r += 1
            r += 1

    for c, w in enumerate([26, 22, 16, 16, 16, 16, 16, 10, 10], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def sheet_items(wb: Workbook, con: sqlite3.Connection, kind: str, title: str, cols: list[str]):
    headers = ["연차", *cols, "기관 표기", "출처"]
    ws, r = _sheet(
        wb, title, headers, [8, 60, 26, 16, 14, 12, 30],
        note="보고서 목록 표에서 건별로 추출한 것이다. 기관 표기는 원본 행에 기관명이 "
             "적혀 있을 때만 채워진다(대부분 논문·발표에는 표기가 없다).",
    )
    for row in con.execute(
        "SELECT * FROM item WHERE kind=? ORDER BY proj_year, date", (kind,)
    ):
        vals = [
            row["proj_year"], row["title"], row["detail"], row["person"], row["date"],
            row["org"], f"{row['doc']} {row['locator']}",
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c, v).alignment = Alignment(wrap_text=(c == 2), vertical="top")
        r += 1


def sheet_achievement(wb: Workbook, con: sqlite3.Connection):
    headers = ["연차", "기관", "세부목표", "연구개발목표", "결과 요약", "달성률", "출처"]
    ws, r = _sheet(wb, "목표달성도", headers, [8, 22, 34, 46, 46, 10, 28])
    for row in con.execute(
        "SELECT * FROM achievement ORDER BY proj_year, org"
    ):
        vals = [
            row["proj_year"], row["org"], row["goal"], row["target_desc"],
            row["result_desc"], row["rate"], f"{row['doc']} {row['locator']}",
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c, v).alignment = Alignment(wrap_text=(c in (3, 4, 5)), vertical="top")
        r += 1


def build() -> Path:
    from ingest import load_config
    from perf_query import connect

    con = connect(load_config())
    wb = Workbook()
    wb.remove(wb.active)

    sheet_indicators(wb, con)
    sheet_quant(wb)
    sheet_items(wb, con, "paper", "논문목록", ["논문명", "학술지명", "주저자", "게재일"])
    sheet_items(wb, con, "patent", "특허목록", ["명칭", "출원/등록번호", "출원인", "일자"])
    sheet_items(wb, con, "presentation", "학술발표", ["발표 제목", "학회·장소", "발표자", "일시"])
    sheet_achievement(wb, con)

    wb.save(OUT)
    con.close()
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"생성: {path.resolve()}")

    from openpyxl import load_workbook

    wb = load_workbook(path)
    for ws in wb.worksheets:
        print(f"  {ws.title:12s} {ws.max_row - 1:4d}행 x {ws.max_column}열")
